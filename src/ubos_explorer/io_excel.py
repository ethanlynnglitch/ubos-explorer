"""Excel access layer.

All engine-specific behaviour (openpyxl for .xlsx, xlrd for legacy .xls) is
confined to this module so that the layout parsers stay engine-independent.

Every workbook is opened read-only. Nothing in this package writes to
``data/raw/``.

Cell values are normalised to exactly one of:

* ``None``          - the cell is empty
* ``float``         - a numeric value (ints are widened to float)
* ``str``           - a text value
* ``CellError``     - an Excel error value such as ``#REF!`` or ``#DIV/0!``
"""

from __future__ import annotations

import hashlib
import re
from dataclasses import dataclass
from pathlib import Path
from typing import Union

import xlrd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

# openpyxl with data_only=True hands back cached error values as plain strings,
# so they have to be recognised by pattern.
_ERROR_TEXT = re.compile(
    r"^#(REF!|DIV/0!|VALUE!|NAME\?|NULL!|NUM!|N/A|GETTING_DATA)$", re.IGNORECASE
)


@dataclass(frozen=True)
class CellError:
    """An Excel error value (``#REF!``, ``#DIV/0!``, ``#VALUE!``, ...)."""

    text: str

    def __str__(self) -> str:  # pragma: no cover - trivial
        return self.text


CellValue = Union[None, float, str, CellError]


def cell_ref(row: int, col: int) -> str:
    """1-based (row, col) -> ``A1``-style reference."""
    return f"{get_column_letter(col)}{row}"


def file_sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with open(path, "rb") as handle:
        for chunk in iter(lambda: handle.read(1 << 20), b""):
            digest.update(chunk)
    return digest.hexdigest()


class Sheet:
    """Read-only view of one worksheet, addressed with 1-based coordinates."""

    def __init__(self, name: str, n_rows: int, n_cols: int):
        self.name = name
        self.n_rows = n_rows
        self.n_cols = n_cols

    def cell(self, row: int, col: int) -> CellValue:  # pragma: no cover
        raise NotImplementedError

    def text(self, row: int, col: int) -> str:
        """Cell as trimmed text; ``''`` for blanks, numbers and error values."""
        value = self.cell(row, col)
        return value.strip() if isinstance(value, str) else ""


class _XlsxSheet(Sheet):
    def __init__(self, worksheet):
        super().__init__(worksheet.title, worksheet.max_row, worksheet.max_column)
        self._ws = worksheet

    def cell(self, row: int, col: int) -> CellValue:
        value = self._ws.cell(row=row, column=col).value
        if value is None:
            return None
        if isinstance(value, bool):
            # Excel booleans are not valid statistical observations.
            return str(value)
        if isinstance(value, (int, float)):
            return float(value)
        text = str(value)
        if _ERROR_TEXT.match(text.strip()):
            return CellError(text.strip())
        return text


class _XlsSheet(Sheet):
    def __init__(self, sheet):
        super().__init__(sheet.name, sheet.nrows, sheet.ncols)
        self._sh = sheet

    def cell(self, row: int, col: int) -> CellValue:
        r, c = row - 1, col - 1
        if r < 0 or c < 0 or r >= self._sh.nrows or c >= self._sh.ncols:
            return None
        kind = self._sh.cell_type(r, c)
        value = self._sh.cell_value(r, c)
        if kind in (xlrd.XL_CELL_EMPTY, xlrd.XL_CELL_BLANK):
            return None
        if kind == xlrd.XL_CELL_NUMBER:
            return float(value)
        if kind == xlrd.XL_CELL_ERROR:
            return CellError(xlrd.error_text_from_code.get(value, f"#ERR{value}"))
        if kind == xlrd.XL_CELL_BOOLEAN:
            return str(bool(value))
        text = str(value)
        if not text.strip():
            return None
        if _ERROR_TEXT.match(text.strip()):
            return CellError(text.strip())
        return text


class Workbook:
    """Engine-agnostic, read-only workbook handle."""

    def __init__(self, path: Path, engine: str):
        self.path = Path(path)
        self.engine = engine
        self._sheets: dict[str, Sheet] = {}
        if engine == "openpyxl":
            self._wb = load_workbook(self.path, data_only=True, read_only=False)
            self.sheet_names = list(self._wb.sheetnames)
        elif engine == "xlrd":
            self._wb = xlrd.open_workbook(str(self.path))
            self.sheet_names = list(self._wb.sheet_names())
        else:
            raise ValueError(f"unsupported excel engine: {engine!r}")

    def sheet(self, name: str) -> Sheet:
        if name not in self._sheets:
            if name not in self.sheet_names:
                raise KeyError(
                    f"worksheet {name!r} not found in {self.path.name}; "
                    f"available: {self.sheet_names}"
                )
            if self.engine == "openpyxl":
                self._sheets[name] = _XlsxSheet(self._wb[name])
            else:
                self._sheets[name] = _XlsSheet(self._wb.sheet_by_name(name))
        return self._sheets[name]

    def close(self) -> None:
        if self.engine == "openpyxl":
            self._wb.close()

    def __enter__(self) -> "Workbook":
        return self

    def __exit__(self, *exc) -> None:
        self.close()


def open_workbook(path: Path, engine: str) -> Workbook:
    return Workbook(path, engine)
